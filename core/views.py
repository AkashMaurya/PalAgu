from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Avg, Q, Sum, F
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
from .models import User, Program, Year, Course, Student, TutorApplication, Session, Feedback, Config, EvaluationYear
from . import forms
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta, date
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def login_view(request):
    """Login page"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        user = authenticate(request, username=email, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.get_full_name()}!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid email or password')
    
    return render(request, 'core/login.html')


def logout_view(request):
    """Logout view"""
    logout(request)
    messages.success(request, 'You have been logged out successfully')
    return redirect('login')


@login_required
def dashboard(request):
    """Main dashboard - routes based on user role"""
    user = request.user
    
    if user.role == 'Admin':
        return redirect('admin_dashboard')
    elif user.role == 'Manager':
        return redirect('manager_analytics')
    elif user.role in ['Student', 'Tutor']:
        return redirect('student_dashboard')
    
    return render(request, 'core/dashboard.html')


@login_required
def admin_dashboard(request):
    """Admin dashboard with stats and quick actions"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')
    
    # Get statistics
    stats = {
        'total_users': User.objects.count(),
        'total_students': User.objects.filter(role='Student').count(),
        'total_tutors': User.objects.filter(role='Tutor').count(),
        'total_sessions': Session.objects.count(),
        'pending_applications': TutorApplication.objects.filter(status='Pending').count(),
    }
    
    # Recent activity
    recent_users = User.objects.all()[:10]
    recent_applications = TutorApplication.objects.all()[:10]
    
    context = {
        'stats': stats,
        'recent_users': recent_users,
        'recent_applications': recent_applications,
    }
    
    return render(request, 'core/admin_dashboard.html', context)


@login_required
def user_management(request):
    """User management page with filtering and search"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')
    
    role_filter = request.GET.get('role', '')
    search = request.GET.get('search', '')
    
    users = User.objects.all()
    
    if role_filter:
        users = users.filter(role=role_filter)
    
    if search:
        users = users.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search) |
            Q(student_id__icontains=search)
        )
    
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'users': page_obj,
        'role_filter': role_filter,
        'search': search,
    }

    return render(request, 'core/user_management.html', context)


@login_required
def create_user(request):
    """Create a new user manually (Admin only)"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    if request.method == 'POST':
        form = forms.AdminUserCreationForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create user
                    user = User.objects.create_user(
                        username=form.cleaned_data['email'],
                        email=form.cleaned_data['email'],
                        password=form.cleaned_data['password'],
                        first_name=form.cleaned_data['first_name'],
                        last_name=form.cleaned_data['last_name'],
                        role=form.cleaned_data['role'],
                        student_id=form.cleaned_data.get('student_id') or None
                    )

                    # If Student role, create Student profile
                    if form.cleaned_data['role'] == 'Student':
                        Student.objects.create(
                            user=user,
                            program=form.cleaned_data['program'],
                            year=form.cleaned_data['year']
                        )

                    messages.success(request, f'User {user.get_full_name()} created successfully!')
                    return redirect('user_management')
            except Exception as e:
                messages.error(request, f'Error creating user: {str(e)}')
    else:
        form = forms.AdminUserCreationForm()

    context = {
        'form': form,
        'programs': Program.objects.all(),
    }

    return render(request, 'core/create_user.html', context)


@login_required
def get_years_by_program(request):
    """HTMX endpoint to get years based on program selection"""
    # HTMX sends the value as a query parameter with the same name as the element
    program_id = request.GET.get('program')

    if not program_id or program_id == '':
        return HttpResponse('<option value="">Select Program First</option>')

    try:
        years = Year.objects.filter(program_id=program_id).order_by('year_number')
        html = '<option value="">Select Year</option>'
        for year in years:
            html += f'<option value="{year.id}">{year.name}</option>'
        return HttpResponse(html)
    except Exception as e:
        return HttpResponse('<option value="">Error loading years</option>')


@login_required
def edit_user(request, user_id):
    """Edit an existing user (Admin only)"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    user_to_edit = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = forms.AdminUserEditForm(request.POST, instance=user_to_edit, user_id=user_id)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f'User {user_to_edit.get_full_name()} updated successfully!')
                return redirect('user_management')
            except Exception as e:
                messages.error(request, f'Error updating user: {str(e)}')
    else:
        form = forms.AdminUserEditForm(instance=user_to_edit, user_id=user_id)

    context = {
        'form': form,
        'user_to_edit': user_to_edit,
    }

    return render(request, 'core/edit_user.html', context)


@login_required
def delete_user(request, user_id):
    """Delete a user (Admin only)"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    user_to_delete = get_object_or_404(User, id=user_id)

    # Prevent deleting yourself
    if user_to_delete.id == request.user.id:
        messages.error(request, 'You cannot delete your own account!')
        return redirect('user_management')

    if request.method == 'POST':
        user_name = user_to_delete.get_full_name()
        user_to_delete.delete()
        messages.success(request, f'User {user_name} deleted successfully!')
        return redirect('user_management')

    context = {
        'user_to_delete': user_to_delete,
    }

    return render(request, 'core/delete_user.html', context)


@login_required
def bulk_upload_users(request):
    """Bulk upload users via XLSX"""
    if request.user.role != 'Admin':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            excel_file = request.FILES['file']
            df = pd.read_excel(excel_file)
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    email = row.get('email')
                    password = row.get('password', 'changeme123')
                    first_name = row.get('first_name')
                    last_name = row.get('last_name')
                    role = row.get('role')
                    student_id = row.get('student_id')
                    program_code = row.get('program')  # New: Program code (MD or NS)
                    year_number = row.get('year')  # New: Year number (1-6 for MD, 1-4 for Nursing)

                    # Validate required fields
                    if not all([email, first_name, last_name, role]):
                        errors.append(f'Row {index + 2}: Missing required fields (email, first_name, last_name, role)')
                        error_count += 1
                        continue

                    # Validate role
                    valid_roles = ['Admin', 'Manager', 'Student', 'Tutor']
                    if role not in valid_roles:
                        errors.append(f'Row {index + 2}: Invalid role "{role}". Must be one of: {", ".join(valid_roles)}')
                        error_count += 1
                        continue

                    # Validate student_id for Student role only
                    if role == 'Student' and (pd.isna(student_id) or not student_id):
                        errors.append(f'Row {index + 2}: Student ID is required for Student role')
                        error_count += 1
                        continue

                    # Validate Program and Year for Student role (PAL Action Plan v2)
                    if role == 'Student':
                        if pd.isna(program_code) or not program_code:
                            errors.append(f'Row {index + 2}: Program is required for Student role')
                            error_count += 1
                            continue

                        if pd.isna(year_number) or not year_number:
                            errors.append(f'Row {index + 2}: Year is required for Student role')
                            error_count += 1
                            continue

                        # Validate program code
                        try:
                            program = Program.objects.get(code=program_code)
                        except Program.DoesNotExist:
                            errors.append(f'Row {index + 2}: Invalid program code "{program_code}". Must be MD or NS')
                            error_count += 1
                            continue

                        # Validate year number based on program
                        try:
                            year_number = int(year_number)
                            if program.code == 'MD' and not (1 <= year_number <= 6):
                                errors.append(f'Row {index + 2}: Year for MD must be between 1 and 6')
                                error_count += 1
                                continue
                            elif program.code == 'NS' and not (1 <= year_number <= 4):
                                errors.append(f'Row {index + 2}: Year for Nursing must be between 1 and 4')
                                error_count += 1
                                continue

                            # Get the Year object
                            year = Year.objects.get(program=program, year_number=year_number)
                        except (ValueError, Year.DoesNotExist):
                            errors.append(f'Row {index + 2}: Invalid year number "{year_number}" for program {program_code}')
                            error_count += 1
                            continue

                    # Check email uniqueness
                    if User.objects.filter(email=email).exists():
                        errors.append(f'Row {index + 2}: Email {email} already exists')
                        error_count += 1
                        continue

                    # Check student_id uniqueness if provided
                    if student_id and not pd.isna(student_id) and User.objects.filter(student_id=student_id).exists():
                        errors.append(f'Row {index + 2}: Student ID {student_id} already exists')
                        error_count += 1
                        continue

                    # Create user
                    with transaction.atomic():
                        user = User.objects.create_user(
                            username=email,
                            email=email,
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                            role=role,
                            student_id=student_id if not pd.isna(student_id) else None
                        )

                        # Create Student profile if role is Student
                        if role == 'Student':
                            Student.objects.create(
                                user=user,
                                program=program,
                                year=year
                            )

                    success_count += 1

                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')
                    error_count += 1
            
            return JsonResponse({
                'success': True,
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def analytics_dashboard(request):
    """Enhanced Analytics dashboard with comprehensive charts and filters"""
    if request.user.role not in ['Admin', 'Manager']:
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    # Get filters from GET params
    selected_program = request.GET.get('program', '')
    selected_year = request.GET.get('year', '')
    selected_course = request.GET.get('course', '')
    selected_tutor = request.GET.get('tutor', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Base queries
    sessions = Session.objects.select_related('tutor', 'learner', 'course', 'evaluation_year')
    feedbacks = Feedback.objects.select_related('tutor', 'learner', 'program', 'year')

    # Apply filters
    if selected_program:
        sessions = sessions.filter(course__program_id=selected_program)
        feedbacks = feedbacks.filter(program_id=selected_program)

    if selected_year:
        sessions = sessions.filter(course__year_id=selected_year)
        feedbacks = feedbacks.filter(year_id=selected_year)

    if selected_course:
        sessions = sessions.filter(course_id=selected_course)

    if selected_tutor:
        sessions = sessions.filter(tutor_id=selected_tutor)
        feedbacks = feedbacks.filter(tutor_id=selected_tutor)

    if start_date:
        sessions = sessions.filter(session_date__gte=start_date)
        feedbacks = feedbacks.filter(session_date__gte=start_date)

    if end_date:
        sessions = sessions.filter(session_date__lte=end_date)
        feedbacks = feedbacks.filter(session_date__lte=end_date)

    # Calculate basic metrics
    total_sessions = sessions.count()
    total_hours = sessions.aggregate(total=Sum('duration'))['total'] or 0
    total_hours = total_hours // 60  # Convert minutes to hours
    total_learners = sessions.values('learner').distinct().count()
    total_feedback = feedbacks.count()
    total_tutors = sessions.values('tutor').distinct().count()
    total_courses = sessions.values('course').distinct().count()
    avg_rating = feedbacks.aggregate(avg=Avg('usefulness_rating'))['avg'] or 0
    total_programs = Program.objects.count()
    total_years = Year.objects.count()

    # Chart data - Sessions by Program
    sessions_by_program = list(sessions.values('course__program__name').annotate(
        count=Count('id')
    ).order_by('-count'))

    # Chart data - Feedback Distribution by Rating
    feedbacks_by_rating = list(feedbacks.values('usefulness_rating').annotate(
        count=Count('id')
    ).order_by('usefulness_rating'))

    # Chart data - Sessions by Status (Pie Chart)
    sessions_by_status = list(sessions.values('status').annotate(
        count=Count('id')
    ))

    # Chart data - Sessions by Course (Top 10)
    sessions_by_course = list(sessions.values('course__name', 'course__code').annotate(
        count=Count('id')
    ).order_by('-count')[:10])

    # Chart data - Top 10 Tutors by Sessions
    top_tutors_sessions = list(sessions.values(
        'tutor__first_name', 'tutor__last_name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10])

    # Chart data - Sessions Over Time (Last 12 months)
    from datetime import datetime, timedelta
    from django.db.models.functions import TruncMonth

    twelve_months_ago = datetime.now() - timedelta(days=365)
    sessions_over_time = list(sessions.filter(
        session_date__gte=twelve_months_ago
    ).annotate(
        month=TruncMonth('session_date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month'))

    # Chart data - Hours by Tutor (Top 10)
    hours_by_tutor = list(sessions.values(
        'tutor__first_name', 'tutor__last_name'
    ).annotate(
        total_hours=Sum('duration')
    ).order_by('-total_hours')[:10])

    # Chart data - Learners by Program (Pie Chart)
    learners_by_program = list(sessions.values('course__program__name').annotate(
        count=Count('learner', distinct=True)
    ).order_by('-count'))

    # Chart data - Feedback Sentiment (Attend Again)
    feedback_attend_again = list(feedbacks.values('attend_again').annotate(
        count=Count('id')
    ))

    # Chart data - Sessions by Day of Week
    from django.db.models.functions import ExtractWeekDay
    sessions_by_weekday = list(sessions.annotate(
        weekday=ExtractWeekDay('session_date')
    ).values('weekday').annotate(
        count=Count('id')
    ).order_by('weekday'))

    # Get filter options
    programs = Program.objects.all()
    years = Year.objects.all()
    courses = Course.objects.all()
    tutors = User.objects.filter(role='Tutor')

    # Prepare metrics dictionary
    metrics = {
        'total_tutors': total_tutors,
        'total_sessions': total_sessions,
        'total_hours': total_hours,
        'total_courses': total_courses,
        'total_learners': total_learners,
        'avg_rating': avg_rating,
        'total_programs': total_programs,
        'total_years': total_years,
        'total_feedback': total_feedback,
    }

    # Get today's date for max attribute
    from datetime import date
    today = date.today().isoformat()

    context = {
        'metrics': metrics,
        'programs': programs,
        'years': years,
        'courses': courses,
        'tutors': tutors,
        'selected_program': selected_program,
        'selected_year': selected_year,
        'selected_course': selected_course,
        'selected_tutor': selected_tutor,
        'start_date': start_date,
        'end_date': end_date,
        'today': today,
        'sessions_by_program': json.dumps(sessions_by_program),
        'feedbacks_by_rating': json.dumps(feedbacks_by_rating),
        'sessions_by_status': json.dumps(sessions_by_status),
        'sessions_by_course': json.dumps(sessions_by_course),
        'top_tutors_sessions': json.dumps(top_tutors_sessions),
        'sessions_over_time': json.dumps(sessions_over_time, default=str),
        'hours_by_tutor': json.dumps(hours_by_tutor),
        'learners_by_program': json.dumps(learners_by_program),
        'feedback_attend_again': json.dumps(feedback_attend_again),
        'sessions_by_weekday': json.dumps(sessions_by_weekday),
        'is_manager': request.user.role == 'Manager',
    }

    return render(request, 'core/analytics_dashboard.html', context)


@login_required
def analytics_export_excel(request):
    """Export analytics data to Excel with filters applied"""
    if request.user.role not in ['Admin', 'Manager']:
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    # Get filters from GET params (same as analytics_dashboard)
    selected_program = request.GET.get('program', '')
    selected_year = request.GET.get('year', '')
    selected_course = request.GET.get('course', '')
    selected_tutor = request.GET.get('tutor', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Base queries
    sessions = Session.objects.select_related('tutor', 'learner', 'course', 'evaluation_year')
    feedbacks = Feedback.objects.select_related('tutor', 'learner', 'program', 'year')

    # Apply filters
    if selected_program:
        sessions = sessions.filter(course__program_id=selected_program)
        feedbacks = feedbacks.filter(program_id=selected_program)

    if selected_year:
        sessions = sessions.filter(course__year_id=selected_year)
        feedbacks = feedbacks.filter(year_id=selected_year)

    if selected_course:
        sessions = sessions.filter(course_id=selected_course)

    if selected_tutor:
        sessions = sessions.filter(tutor_id=selected_tutor)
        feedbacks = feedbacks.filter(tutor_id=selected_tutor)

    if start_date:
        sessions = sessions.filter(session_date__gte=start_date)
        feedbacks = feedbacks.filter(session_date__gte=start_date)

    if end_date:
        sessions = sessions.filter(session_date__lte=end_date)
        feedbacks = feedbacks.filter(session_date__lte=end_date)

    # Create workbook
    wb = Workbook()

    # Define styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Summary Metrics
    ws_summary = wb.active
    ws_summary.title = "Summary Metrics"

    # Calculate metrics
    total_sessions = sessions.count()
    total_hours = sessions.aggregate(total=Sum('duration'))['total'] or 0
    total_hours = total_hours // 60
    total_learners = sessions.values('learner').distinct().count()
    total_feedback = feedbacks.count()
    total_tutors = sessions.values('tutor').distinct().count()
    total_courses = sessions.values('course').distinct().count()
    avg_rating = feedbacks.aggregate(avg=Avg('usefulness_rating'))['avg'] or 0

    # Add summary data
    summary_data = [
        ["Metric", "Value"],
        ["Total Sessions", total_sessions],
        ["Total Hours", total_hours],
        ["Total Learners", total_learners],
        ["Total Tutors", total_tutors],
        ["Total Courses", total_courses],
        ["Total Feedback", total_feedback],
        ["Average Rating", f"{avg_rating:.2f}"],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

    # Auto-size columns
    for col in range(1, 3):
        ws_summary.column_dimensions[get_column_letter(col)].width = 25

    # Sheet 2: Sessions Detail
    ws_sessions = wb.create_sheet("Sessions Detail")

    session_headers = ["ID", "Date", "Tutor", "Learner", "Course", "Program", "Duration (min)", "Status"]
    for col_idx, header in enumerate(session_headers, 1):
        cell = ws_sessions.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, session in enumerate(sessions.order_by('-session_date'), 2):
        session_data = [
            session.id,
            session.session_date.strftime('%Y-%m-%d') if session.session_date else '',
            session.tutor.get_full_name() if session.tutor else '',
            session.learner.get_full_name() if session.learner else '',
            f"{session.course.code} - {session.course.name}" if session.course else '',
            session.course.program.name if session.course and session.course.program else '',
            session.duration or 0,
            session.status or '',
        ]
        for col_idx, value in enumerate(session_data, 1):
            cell = ws_sessions.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Auto-size columns
    for col in range(1, len(session_headers) + 1):
        ws_sessions.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 3: Feedback Detail
    ws_feedback = wb.create_sheet("Feedback Detail")

    feedback_headers = ["ID", "Date", "Tutor", "Learner", "Program", "Year", "Explanation Rating", "Usefulness Rating", "Attend Again", "Well Organized"]
    for col_idx, header in enumerate(feedback_headers, 1):
        cell = ws_feedback.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row_idx, feedback in enumerate(feedbacks.order_by('-session_date'), 2):
        feedback_data = [
            feedback.id,
            feedback.session_date.strftime('%Y-%m-%d') if feedback.session_date else '',
            feedback.tutor.get_full_name() if feedback.tutor else '',
            feedback.learner.get_full_name() if feedback.learner else '',
            feedback.program.name if feedback.program else '',
            feedback.year.name if feedback.year else '',
            feedback.explanation_rating or '',
            feedback.usefulness_rating or '',
            'Yes' if feedback.attend_again else 'No',
            'Yes' if feedback.well_organized else 'No',
        ]
        for col_idx, value in enumerate(feedback_data, 1):
            cell = ws_feedback.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Auto-size columns
    for col in range(1, len(feedback_headers) + 1):
        ws_feedback.column_dimensions[get_column_letter(col)].width = 18

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Generate filename with timestamp and filters
    filename_parts = ['PAL_Analytics']
    if selected_program:
        try:
            program = Program.objects.get(id=selected_program)
            filename_parts.append(program.code)
        except:
            pass
    if start_date:
        filename_parts.append(f"from_{start_date}")
    if end_date:
        filename_parts.append(f"to_{end_date}")

    filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
    filename = '_'.join(filename_parts) + '.xlsx'

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def analytics_export_pdf(request):
    """Export analytics dashboard to PDF (PAL Action Plan v2)"""
    if request.user.role not in ['Admin', 'Manager']:
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    try:
        from weasyprint import HTML, CSS
        from django.template.loader import render_to_string

        # Get the same context as analytics_dashboard
        # Reuse the analytics logic
        filter_form = forms.AnalyticsFilterForm(request.GET or None)

        sessions = Session.objects.select_related('tutor', 'learner', 'course', 'evaluation_year')
        feedbacks = Feedback.objects.select_related('tutor', 'learner', 'program', 'year')

        # Apply filters
        if filter_form.is_valid():
            evaluation_year = filter_form.cleaned_data.get('evaluation_year')
            program = filter_form.cleaned_data.get('program')
            start_date = filter_form.cleaned_data.get('start_date')
            end_date = filter_form.cleaned_data.get('end_date')

            if evaluation_year:
                sessions = sessions.filter(evaluation_year=evaluation_year)
                feedbacks = feedbacks.filter(session_date__range=[evaluation_year.start_date, evaluation_year.end_date])

            if program:
                sessions = sessions.filter(course__program=program)
                feedbacks = feedbacks.filter(program=program)

            if start_date:
                sessions = sessions.filter(session_date__gte=start_date)
                feedbacks = feedbacks.filter(session_date__gte=start_date)

            if end_date:
                sessions = sessions.filter(session_date__lte=end_date)
                feedbacks = feedbacks.filter(session_date__lte=end_date)

        # Calculate metrics
        total_sessions = sessions.count()
        total_hours = (sessions.aggregate(total=Sum('duration'))['total'] or 0) // 60
        total_learners = sessions.values('learner').distinct().count()
        total_feedback = feedbacks.count()

        top_tutors_by_sessions = sessions.values('tutor__first_name', 'tutor__last_name').annotate(
            session_count=Count('id')
        ).order_by('-session_count')[:5]

        top_rated_tutors = feedbacks.values('tutor__first_name', 'tutor__last_name').annotate(
            avg_rating=Avg('usefulness_rating'),
            feedback_count=Count('id')
        ).filter(feedback_count__gte=3).order_by('-avg_rating')[:5]

        trendy_topics = feedbacks.values('topic').annotate(
            count=Count('id')
        ).order_by('-count')[:10]

        busy_courses = sessions.values('course__name', 'course__code').annotate(
            session_count=Count('id')
        ).order_by('-session_count')[:10]

        context = {
            'total_sessions': total_sessions,
            'total_hours': total_hours,
            'total_learners': total_learners,
            'total_feedback': total_feedback,
            'total_tutors': User.objects.filter(role='Tutor').count(),
            'top_tutors_by_sessions': top_tutors_by_sessions,
            'top_rated_tutors': top_rated_tutors,
            'trendy_topics': trendy_topics,
            'busy_courses': busy_courses,
            'generated_date': timezone.now(),
            'filters': filter_form.cleaned_data if filter_form.is_valid() else {},
        }

        # Render HTML template
        html_string = render_to_string('core/analytics_pdf.html', context)

        # Generate PDF
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf_file = html.write_pdf()

        # Create response
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="pal_analytics_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

        return response

    except ImportError:
        messages.error(request, 'WeasyPrint is not installed. Please install it to export PDFs.')
        return redirect('analytics')
    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('analytics')


@login_required
def manager_analytics(request):
    """Manager analytics view (read-only)"""
    if request.user.role != 'Manager':
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    return analytics_dashboard(request)


@login_required
def student_dashboard(request):
    """Student/Tutor dashboard"""
    return render(request, 'core/student_dashboard.html')


@login_required
def settings_view(request):
    """Admin settings page"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')
    
    if request.method == 'POST':
        for key in request.POST:
            if key != 'csrfmiddlewaretoken':
                config, created = Config.objects.get_or_create(key=key)
                config.value = request.POST[key]
                config.save()
        
        messages.success(request, 'Settings updated successfully')
        return redirect('settings')
    
    configs = {c.key: c.value for c in Config.objects.all()}

    return render(request, 'core/settings.html', {'configs': configs})


@login_required
def manage_evaluation_years(request):
    """Manage Evaluation Years (PAL Action Plan v2)"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    evaluation_years = EvaluationYear.objects.all().order_by('-start_date')

    return render(request, 'core/manage_evaluation_years.html', {
        'evaluation_years': evaluation_years
    })


@login_required
def create_evaluation_year(request):
    """Create new Evaluation Year"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    if request.method == 'POST':
        form = forms.EvaluationYearForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Evaluation year created successfully!')
            return redirect('manage_evaluation_years')
    else:
        form = forms.EvaluationYearForm()

    return render(request, 'core/evaluation_year_form.html', {
        'form': form,
        'action': 'Create'
    })


@login_required
def edit_evaluation_year(request, year_id):
    """Edit Evaluation Year"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    evaluation_year = get_object_or_404(EvaluationYear, id=year_id)

    if request.method == 'POST':
        form = forms.EvaluationYearForm(request.POST, instance=evaluation_year)
        if form.is_valid():
            form.save()
            messages.success(request, 'Evaluation year updated successfully!')
            return redirect('manage_evaluation_years')
    else:
        form = forms.EvaluationYearForm(instance=evaluation_year)

    return render(request, 'core/evaluation_year_form.html', {
        'form': form,
        'action': 'Edit',
        'evaluation_year': evaluation_year
    })


@login_required
def delete_evaluation_year(request, year_id):
    """Delete Evaluation Year"""
    if request.user.role != 'Admin':
        messages.error(request, 'Access denied')
        return redirect('dashboard')

    evaluation_year = get_object_or_404(EvaluationYear, id=year_id)

    if request.method == 'POST':
        evaluation_year.delete()
        messages.success(request, 'Evaluation year deleted successfully!')
        return redirect('manage_evaluation_years')

    return render(request, 'core/confirm_delete.html', {
        'object': evaluation_year,
        'object_type': 'Evaluation Year'
    })


# Multi-step Registration Views

@login_required
def student_registration_start(request):
    """Student registration - Step 1: Personal Information"""
    if request.method == 'POST':
        # Pass the user instance to the form
        form = forms.StudentRegistrationForm(request.POST, instance=request.user, user=request.user)
        if form.is_valid():
            # Store data in session (use existing user data if not changed)
            request.session['student_reg_step1'] = {
                'user_id': request.user.id,  # Store user ID to update existing user
                'email': form.cleaned_data['email'],
                'first_name': form.cleaned_data['first_name'],
                'last_name': form.cleaned_data['last_name'],
                'student_id': form.cleaned_data['student_id'],
            }
            return JsonResponse({'next_step': 'step2'})
        else:
            return JsonResponse({'errors': form.errors}, status=400)

    # Pre-fill form with logged-in user's instance
    form = forms.StudentRegistrationForm(instance=request.user, user=request.user)
    return render(request, 'core/student_registration_step1.html', {'form': form})


@login_required
def student_registration_step2(request):
    """Student registration - Step 2: Program Selection (Pre-filled from user's Student profile)"""
    if 'student_reg_step1' not in request.session:
        return redirect('student_registration_start')

    # Check if user has a Student profile with program and year
    initial_data = {}
    student_profile = None
    try:
        student_profile = request.user.student_profile
        if student_profile.program and student_profile.year:
            initial_data = {
                'program': student_profile.program,
                'year': student_profile.year,
            }
    except Student.DoesNotExist:
        pass  # No profile, show empty form

    if request.method == 'POST':
        form = forms.StudentProgramForm(request.POST)
        if form.is_valid():
            request.session['student_reg_step2'] = {
                'program_id': form.cleaned_data['program'].id,
                'year_id': form.cleaned_data['year'].id,
            }
            return JsonResponse({'next_step': 'step3'})
        else:
            return JsonResponse({'errors': form.errors}, status=400)

    # Pre-fill form with student profile data if available
    if initial_data:
        form = forms.StudentProgramForm(initial=initial_data)
    else:
        form = forms.StudentProgramForm()

    programs = Program.objects.all()
    return render(request, 'core/student_registration_step2.html', {
        'form': form,
        'programs': programs,
        'student_profile': student_profile,
    })


@login_required
def student_registration_step3(request):
    """Student registration - Step 3: Course Selection (Cumulative course visibility)"""
    if 'student_reg_step2' not in request.session:
        return redirect('student_registration_start')

    year_id = request.session['student_reg_step2']['year_id']
    program_id = request.session['student_reg_step2']['program_id']

    # Get year_number for cumulative course visibility
    year = Year.objects.get(id=year_id)
    year_number = year.year_number

    # Debug: Check if courses exist for this program
    courses_count = Course.objects.filter(
        program_id=program_id,
        year__year_number__lte=year_number
    ).count()
    print(f"DEBUG: Year ID: {year_id}, Year Number: {year_number}, Program ID: {program_id}, Cumulative courses: {courses_count}")

    max_selections = int(Config.objects.get(key='maxCourseSelections').value)

    if request.method == 'POST':
        form = forms.StudentCourseSelectionForm(
            request.POST,
            program_id=program_id,
            year_number=year_number,
            max_selections=max_selections
        )
        if form.is_valid():
            course_ids = [c.id for c in form.cleaned_data['courses']]
            request.session['student_reg_step3'] = {
                'course_ids': course_ids
            }
            
            # Complete registration
            try:
                step1 = request.session.get('student_reg_step1')
                step2 = request.session.get('student_reg_step2')

                if not step1 or not step2:
                    return JsonResponse({'error': 'Registration data missing. Please start over.'}, status=400)

                # Update existing user or create new one
                if 'user_id' in step1:
                    # Update existing logged-in user
                    user = User.objects.get(id=step1['user_id'])
                    user.email = step1['email']
                    user.first_name = step1['first_name']
                    user.last_name = step1['last_name']
                    user.student_id = step1['student_id']
                    user.role = 'Student'
                    user.save()
                else:
                    # Create new user (fallback for non-logged-in users)
                    user = User.objects.create_user(
                        username=step1['email'],
                        email=step1['email'],
                        password=step1.get('password', 'changeme123'),
                        first_name=step1['first_name'],
                        last_name=step1['last_name'],
                        role='Student',
                        student_id=step1['student_id']
                    )

                # Create or update student profile
                student, created = Student.objects.update_or_create(
                    user=user,
                    defaults={
                        'program_id': step2['program_id'],
                        'year_id': step2['year_id'],
                    }
                )

                # Note: Selected courses are stored in session but not persisted to database
                # They can be used for matching with tutors or other purposes

                action = "updated" if not created else "created"

                # Clear session data
                if 'student_reg_step1' in request.session:
                    del request.session['student_reg_step1']
                if 'student_reg_step2' in request.session:
                    del request.session['student_reg_step2']
                if 'student_reg_step3' in request.session:
                    del request.session['student_reg_step3']

                messages.success(request, f'Student profile {action} successfully! Welcome to the PAL Program.')
                return JsonResponse({
                    'success': True,
                    'redirect': '/student/',
                    'message': f'Student profile {action} successfully!'
                })

            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"Registration error: {error_details}")  # Log to console
                return JsonResponse({'error': f'Registration failed: {str(e)}'}, status=400)
        else:
            return JsonResponse({'errors': form.errors}, status=400)

    form = forms.StudentCourseSelectionForm(
        program_id=program_id,
        year_number=year_number,
        max_selections=max_selections
    )

    # Get program and year info for display
    try:
        program = Program.objects.get(id=program_id)
        # year already fetched above

        # Debug: Print form queryset
        print(f"DEBUG: Form queryset count: {form.fields['courses'].queryset.count()}")
        print(f"DEBUG: Form queryset: {list(form.fields['courses'].queryset.values_list('code', 'name', 'year__year_number'))}")
    except Exception as e:
        print(f"DEBUG ERROR: {e}")
        program = None

    return render(request, 'core/student_registration_step3.html', {
        'form': form,
        'max_selections': max_selections,
        'program': program,
        'year': year,
        'courses_count': courses_count,
        'debug_year_id': year_id,
        'debug_program_id': program_id
    })


def tutor_registration_start(request):
    """Tutor registration - Step 1: Personal and Interest Information (PAL Action Plan v2)"""
    # Pre-fill form with logged-in user data
    initial_data = {}
    if request.user.is_authenticated:
        initial_data = {
            'name': request.user.get_full_name(),
            'student_id': request.user.student_id or '',
            'email': request.user.email,
        }

    if request.method == 'POST':
        form = forms.TutorApplicationStep1Form(request.POST)
        if form.is_valid():
            # Check if interested_as_tutor is True
            if not form.cleaned_data.get('interested_as_tutor'):
                # User is not interested, show thank you message
                return JsonResponse({
                    'not_interested': True,
                    'message': 'We thank you for your feedback.'
                })

            # Store step 1 data in session
            request.session['tutor_reg_step1'] = {
                'name': form.cleaned_data['name'],
                'student_id': form.cleaned_data['student_id'],
                'email': form.cleaned_data['email'],
                'mobile': form.cleaned_data.get('mobile', ''),
                'engaged_in_pal': form.cleaned_data['engaged_in_pal'],
                'wants_training': form.cleaned_data.get('wants_training', False),
                'wants_certificate': form.cleaned_data.get('wants_certificate', False),
                'suggestions': form.cleaned_data.get('suggestions', ''),
                'interested_as_tutor': form.cleaned_data['interested_as_tutor'],
            }
            return JsonResponse({'next_step': 'step2'})
        else:
            return JsonResponse({'errors': form.errors}, status=400)

    form = forms.TutorApplicationStep1Form(initial=initial_data)
    return render(request, 'core/tutor_registration_step1.html', {
        'form': form,
        'user': request.user
    })


def tutor_registration_step2(request):
    """Tutor registration - Step 2: Academic Details (PAL Action Plan v2)"""
    if 'tutor_reg_step1' not in request.session:
        return redirect('tutor_registration_start')

    # Pre-fill form with student profile data if available
    initial_data = {}
    student_profile = None

    if request.user.is_authenticated:
        try:
            student_profile = request.user.student_profile
            # Pre-fill with student's program and year, but allow them to change it
            initial_data = {
                'program': student_profile.program,
                'year': student_profile.year,
            }
        except Student.DoesNotExist:
            pass  # No profile, show empty form

    if request.method == 'POST':
        form = forms.TutorApplicationStep2Form(request.POST)
        print(f"DEBUG: Form is_valid: {form.is_valid()}")
        print(f"DEBUG: Form errors: {form.errors}")
        print(f"DEBUG: POST data: {request.POST}")
        if form.is_valid():
            # Validate GPA against minimum requirement (only if provided)
            gpa = form.cleaned_data.get('gpa')
            if gpa is not None:
                try:
                    min_gpa_config = Config.objects.get(key='minGpaForTutor')
                    min_gpa = float(min_gpa_config.value)
                except (Config.DoesNotExist, ValueError):
                    min_gpa = 3.0  # Default minimum GPA

                if gpa < min_gpa:
                    return JsonResponse({
                        'error': f'Minimum GPA of {min_gpa} required to become a tutor'
                    }, status=400)

            # Check consent
            if not form.cleaned_data.get('consent'):
                return JsonResponse({
                    'error': 'You must consent to the terms to proceed'
                }, status=400)

            # Get checkbox values (multiple selections) directly from POST data
            # and convert to comma-separated strings
            preferred_days = request.POST.getlist('preferred_days')
            preferred_times = request.POST.getlist('preferred_times')

            # Store step 2 data in session
            request.session['tutor_reg_step2'] = {
                'program_id': form.cleaned_data['program'].id,
                'year_id': form.cleaned_data['year'].id,
                'year_number': form.cleaned_data['year'].year_number,
                'gpa': float(gpa) if gpa is not None else None,
                'motivation': form.cleaned_data['motivation'],
                'confidence_rating': int(form.cleaned_data['confidence_rating']),
                'preferred_days': ', '.join(preferred_days) if preferred_days else '',
                'preferred_times': ', '.join(preferred_times) if preferred_times else '',
                'preferred_mode': form.cleaned_data.get('preferred_mode', ''),
                'max_sessions_per_week': form.cleaned_data.get('max_sessions_per_week'),
                'consent': form.cleaned_data['consent'],
            }
            return JsonResponse({'next_step': 'step3'})
        else:
            print(f"DEBUG: Returning form errors: {dict(form.errors)}")
            return JsonResponse({'errors': dict(form.errors)}, status=400)

    form = forms.TutorApplicationStep2Form(initial=initial_data)
    programs = Program.objects.all()
    return render(request, 'core/tutor_registration_step2.html', {
        'form': form,
        'programs': programs,
        'student_profile': student_profile,
    })


def tutor_registration_step3(request):
    """Tutor registration - Step 3: Course Selection (PAL Action Plan v2)"""
    if 'tutor_reg_step2' not in request.session:
        return redirect('tutor_registration_start')

    step2 = request.session['tutor_reg_step2']
    program_id = step2['program_id']
    year_number = step2['year_number']

    if request.method == 'POST':
        form = forms.TutorApplicationStep3Form(
            request.POST,
            program_id=program_id,
            year_number=year_number
        )
        print(f"DEBUG Step3: Form is_valid: {form.is_valid()}")
        print(f"DEBUG Step3: Form errors: {form.errors}")
        print(f"DEBUG Step3: POST data: {request.POST}")
        if form.is_valid():
            courses = form.cleaned_data['courses']
            print(f"DEBUG Step3: Courses selected: {courses}, count: {len(courses)}")

            # Validate exactly 3 courses
            if len(courses) != 3:
                print(f"DEBUG Step3: Course count validation failed: {len(courses)} != 3")
                return JsonResponse({
                    'error': 'You must select exactly 3 courses'
                }, status=400)

            # Save application
            try:
                with transaction.atomic():
                    step1 = request.session['tutor_reg_step1']
                    print(f"DEBUG Step3: Step1 data: {step1}")

                    # Check if user exists or create new one
                    user, created = User.objects.get_or_create(
                        student_id=step1['student_id'],
                        defaults={
                            'username': step1['email'],
                            'email': step1['email'],
                            'first_name': step1['name'].split()[0] if step1['name'] else '',
                            'last_name': ' '.join(step1['name'].split()[1:]) if len(step1['name'].split()) > 1 else '',
                            'role': 'Student',  # Default role, can be updated to Tutor upon approval
                        }
                    )

                    # If user exists, update email if different
                    if not created and user.email != step1['email']:
                        user.email = step1['email']
                        user.username = step1['email']
                        user.save()

                    # Create tutor application
                    tutor_app = TutorApplication.objects.create(
                        user=user,
                        mobile=step1.get('mobile', ''),
                        engaged_in_pal=step1['engaged_in_pal'],
                        wants_training=step1.get('wants_training', False),
                        wants_certificate=step1.get('wants_certificate', False),
                        suggestions=step1.get('suggestions', ''),
                        interested_as_tutor=step1['interested_as_tutor'],
                        program_id=program_id,
                        year_id=step2['year_id'],
                        gpa=step2['gpa'],
                        motivation=step2['motivation'],
                        confidence_rating=step2['confidence_rating'],
                        preferred_days=step2.get('preferred_days', ''),
                        preferred_times=step2.get('preferred_times', ''),
                        preferred_mode=step2.get('preferred_mode', ''),
                        max_sessions_per_week=step2.get('max_sessions_per_week'),
                        consent=step2['consent'],
                        status='Approved'  # Auto-approve tutors upon registration
                    )

                    # Add courses (exactly 3)
                    tutor_app.courses.set([c.id for c in courses])

                    # Clear session data
                    del request.session['tutor_reg_step1']
                    del request.session['tutor_reg_step2']

                    return JsonResponse({
                        'success': True,
                        'message': "Congratulations! Your tutor application has been approved. You can now start tutoring sessions and students can submit feedback for you.",
                        'redirect': '/dashboard/'
                    })

            except Exception as e:
                import traceback
                print(f"DEBUG Step3: Exception occurred: {str(e)}")
                print(f"DEBUG Step3: Traceback: {traceback.format_exc()}")
                return JsonResponse({'error': str(e)}, status=400)
        else:
            return JsonResponse({'errors': form.errors}, status=400)

    form = forms.TutorApplicationStep3Form(
        program_id=program_id,
        year_number=year_number
    )

    # Get program and year for display
    program = Program.objects.get(id=program_id)
    year = Year.objects.get(id=step2['year_id'])

    return render(request, 'core/tutor_registration_step3.html', {
        'form': form,
        'program': program,
        'year': year,
        'step1': request.session['tutor_reg_step1'],
        'step2': step2
    })


# Step 4 removed - now using 3-step process as per PAL Action Plan v2


# Enhanced Student/Tutor Dashboards

@login_required
def enhanced_student_dashboard(request):
    """Enhanced student dashboard with sessions and feedback"""
    user = request.user
    
    if user.role == 'Student':
        # Get student profile
        try:
            student = user.student_profile
        except:
            student = None
        
        # Get sessions as learner
        sessions = Session.objects.filter(learner=user).select_related('tutor', 'course')
        
        # Get pending feedback
        pending_feedback = sessions.filter(status='Completed').exclude(
            id__in=Feedback.objects.values_list('session_id', flat=True)
        )
        
        context = {
            'student': student,
            'sessions': sessions,
            'pending_feedback': pending_feedback,
            'upcoming_sessions': sessions.filter(status='Scheduled').order_by('session_date')[:5],
            'completed_sessions': sessions.filter(status='Completed').count(),
        }
        
    elif user.role == 'Tutor':
        # Get tutor applications
        tutor_apps = TutorApplication.objects.filter(user=user).prefetch_related('courses')
        approved_apps = tutor_apps.filter(status='Approved')
        
        # Get sessions as tutor
        sessions = Session.objects.filter(tutor=user).select_related('learner', 'course')
        
        context = {
            'tutor_apps': tutor_apps,
            'approved_apps': approved_apps,
            'sessions': sessions,
            'upcoming_sessions': sessions.filter(status='Scheduled').order_by('session_date')[:5],
            'total_hours': sessions.filter(status='Completed').aggregate(
                total=Sum('duration')
            )['total'] or 0,
            'total_sessions': sessions.filter(status='Completed').count(),
        }
    else:
        context = {}
    
    return render(request, 'core/enhanced_student_dashboard.html', context)


@login_required
def create_session(request):
    """Create a new tutoring session"""
    if request.user.role != 'Tutor':
        messages.error(request, 'Only tutors can create sessions')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = forms.SessionCreateForm(request.POST, tutor=request.user)
        if form.is_valid():
            session = form.save(commit=False)
            session.tutor = request.user
            session.save()
            messages.success(request, 'Session created successfully!')
            return redirect('enhanced_student_dashboard')
    else:
        form = forms.SessionCreateForm(tutor=request.user)
    
    return render(request, 'core/create_session.html', {'form': form})


@login_required
def submit_feedback(request, session_id):
    """Submit feedback for a completed session"""
    session = get_object_or_404(Session, id=session_id, learner=request.user, status='Completed')
    
    # Check if feedback already exists
    if hasattr(session, 'feedback'):
        messages.warning(request, 'Feedback already submitted for this session')
        return redirect('enhanced_student_dashboard')
    
    if request.method == 'POST':
        form = forms.SessionFeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.session = session
            feedback.save()
            messages.success(request, 'Feedback submitted successfully!')
            return redirect('enhanced_student_dashboard')
    else:
        form = forms.SessionFeedbackForm()
    
    return render(request, 'core/submit_feedback.html', {
        'form': form,
        'session': session
    })


@login_required
def learner_feedback_submit(request):
    """New Learner Feedback Form (PAL Action Plan v2)"""
    # Must be a student
    if request.user.role != 'Student':
        messages.error(request, 'Only students can submit feedback')
        return redirect('dashboard')

    # Get student profile (optional - will show warning if missing)
    student = None
    has_profile = False
    try:
        student = request.user.student_profile
        has_profile = True
    except Student.DoesNotExist:
        # Allow submission without profile, but show warning
        messages.warning(request, 'Note: You have not completed your student registration. Please complete it to get personalized tutor recommendations.')

    if request.method == 'POST':
        form = forms.LearnerFeedbackForm(request.POST, learner=request.user)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.learner = request.user

            # Program and year are now selected by the student in the form
            # No need to auto-assign them

            # session_date is auto-set by model (auto_now_add=True)
            feedback.save()

            messages.success(request, 'Thank you for your feedback!')
            return redirect('student_dashboard')
    else:
        form = forms.LearnerFeedbackForm(learner=request.user)

    return render(request, 'core/learner_feedback.html', {
        'form': form,
        'student': student,
        'has_profile': has_profile
    })


@login_required
def get_years_for_program(request):
    """AJAX endpoint to get years for a selected program

    For tutor registration: Returns only years below the student's current year
    (near-peer tutoring - e.g., Year 4 student can tutor Years 1, 2, 3)
    """
    program_id = request.GET.get('program_id')
    if program_id:
        # Check if user has a student profile to filter years
        try:
            student_profile = request.user.student_profile
            # Only return years below student's current year
            years = Year.objects.filter(
                program_id=program_id,
                year_number__lt=student_profile.year.year_number
            ).values('id', 'name', 'year_number').order_by('year_number')
        except Student.DoesNotExist:
            # No student profile, return all years for the program
            years = Year.objects.filter(program_id=program_id).values('id', 'name', 'year_number').order_by('year_number')

        return JsonResponse({'years': list(years)})
    return JsonResponse({'years': []})
